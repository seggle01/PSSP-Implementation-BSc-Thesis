import subprocess
import tempfile
import shutil
import pandas as pd
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

######### Fixed Column Orders #########
Q3_COLUMN_ORDER = [
    'Accuracy',
    'SOV_99',    'SOV_refine',
    'SOV_99_H',  'SOV_99_E',  'SOV_99_C',
    'SOV_refine_H', 'SOV_refine_E', 'SOV_refine_C',
]

Q8_COLUMN_ORDER = [
    'Accuracy',
    'SOV_99',    'SOV_refine',
    'SOV_99_G',  'SOV_99_H',  'SOV_99_I',  'SOV_99_B',
    'SOV_99_E',  'SOV_99_S',  'SOV_99_T',  'SOV_99_C',
    'SOV_refine_G', 'SOV_refine_H', 'SOV_refine_I', 'SOV_refine_B',
    'SOV_refine_E', 'SOV_refine_S', 'SOV_refine_T', 'SOV_refine_C',
]

Q3_CLASS_ORDER = ['H', 'E', 'C']
Q8_CLASS_ORDER = ['G', 'H', 'I', 'B', 'E', 'S', 'T', 'C']

######### Color Palette #########
HEADER_MODEL = PatternFill("solid", fgColor="2E4057")
HEADER_MAIN  = PatternFill("solid", fgColor="048A81")
BEST_FILL    = PatternFill("solid", fgColor="FFD700")
WHITE_FONT   = Font(color="FFFFFF", bold=True)
DARK_FONT    = Font(color="1A1A2E", bold=False)
CENTER       = Alignment(horizontal="center", vertical="center")
thin_border  = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

MODEL_COLOR_PALETTE = ["EAF4FB", "FFF8EA", "F0EAF8", "EAFAF1", "FEF9E7"]

CLASS_COLORS = [
    PatternFill("solid", fgColor="54C6EB"),
    PatternFill("solid", fgColor="8EE3EF"),
    PatternFill("solid", fgColor="CAF0F8"),
    PatternFill("solid", fgColor="B5EAD7"),
    PatternFill("solid", fgColor="FFDAC1"),
    PatternFill("solid", fgColor="E2F0CB"),
    PatternFill("solid", fgColor="FFB7B2"),
    PatternFill("solid", fgColor="C7CEEA"),
]

######### Dynamic Header Fill #########
def _get_col_fill(col_name, class_order):
    if col_name == 'Model':
        return HEADER_MODEL
    if col_name in ('Accuracy', 'SOV_99', 'SOV_refine'):
        return HEADER_MAIN
    cls = col_name.split('_')[-1]
    idx = class_order.index(cls) if cls in class_order else 0
    return CLASS_COLORS[idx % len(CLASS_COLORS)]

######### Pair discovery #########
def find_fasta_pairs(fasta_dir):
    files    = [f for f in os.listdir(fasta_dir) if f.endswith('.fasta')]
    pred_map = {
        f[:-len('_pred.fasta')]: os.path.abspath(os.path.join(fasta_dir, f))
        for f in files if f.endswith('_pred.fasta')
    }
    true_map = {
        f[:-len('_true.fasta')]: os.path.abspath(os.path.join(fasta_dir, f))
        for f in files if f.endswith('_true.fasta')
    }

    paired = {}
    for base in pred_map:
        if base in true_map:
            paired[base] = {'pred': pred_map[base], 'true': true_map[base]}
        else:
            print(f"No true file for: {pred_map[base]}")
    for base in true_map:
        if base not in pred_map:
            print(f"No pred file for: {true_map[base]}")

    print(f"{len(paired)} valid pair(s) found in: {fasta_dir}\n")
    return paired

######### Parse SOV output #########
def parse_sov_output(output_text):
    lines   = output_text.strip().split('\n')
    results = {}
    for line in lines:
        line = line.strip()
        if line.startswith('SOV_99_i'):
            parts = line.split()
            results[f'SOV_99_{parts[1]}'] = float(parts[2])
        elif line.startswith('SOV_refine_i'):
            parts = line.split()
            results[f'SOV_refine_{parts[1]}'] = float(parts[2])
        elif line.startswith('Accuracy'):
            results['Accuracy'] = float(line.split()[1])
        elif line.startswith('SOV_99') and 'SOV_99_i' not in line:
            results['SOV_99'] = float(line.split()[1])
        elif line.startswith('SOV_refine') and 'SOV_refine_i' not in line:
            results['SOV_refine'] = float(line.split()[1])
    return results

######### Run one pair #########
def run_one_pair(model_name, pred_path, true_path, perl_script_path):
    perl_dir = os.path.dirname(os.path.abspath(perl_script_path))

    with tempfile.TemporaryDirectory() as tmpdir:
        shutil.copy2(true_path, os.path.join(tmpdir, 'true.fasta'))
        shutil.copy2(pred_path, os.path.join(tmpdir, 'pred.fasta'))

        for f in os.listdir(perl_dir):
            src = os.path.join(perl_dir, f)
            if os.path.isfile(src):
                shutil.copy2(src, os.path.join(tmpdir, f))

        cmd = ['perl', 'SOV_refine.pl', 'true.fasta', 'pred.fasta', '0.5']

        try:
            result = subprocess.run(
                cmd, capture_output=True, text=True, check=True, cwd=tmpdir)
            output = result.stdout
        except subprocess.CalledProcessError as e:
            print(f"  [{model_name}] Perl error — {e.stderr}")
            return model_name, None

    results = parse_sov_output(output)
    print(f"  [{model_name}] done — "
          f"Acc: {results['Accuracy']:.4f} | "
          f"SOV_99: {results['SOV_99']:.4f} | "
          f"SOV_refine: {results['SOV_refine']:.4f}")
    return model_name, results

######### Excel helpers #########
def _style_header_row(ws, cols, class_order):
    for col_idx, col_name in enumerate(cols, start=1):
        cell           = ws.cell(row=1, column=col_idx)
        cell.fill      = _get_col_fill(col_name, class_order)
        cell.font      = WHITE_FONT
        cell.alignment = CENTER
        cell.border    = thin_border
    ws.row_dimensions[1].height = 25


def _auto_col_widths(ws, cols, num_data_rows):
    for col_idx, col_name in enumerate(cols, start=1):
        max_len = len(str(col_name))
        for r in range(2, num_data_rows + 2):
            val      = ws.cell(row=r, column=col_idx).value
            cell_str = f"{val:.2f}" if isinstance(val, float) else str(val or '')
            max_len  = max(max_len, len(cell_str))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 30)


def _apply_number_format(ws, cols, metric_cols, num_data_rows):
    for col_idx, col_name in enumerate(cols, start=1):
        if col_name in metric_cols:
            for r in range(2, num_data_rows + 2):
                cell = ws.cell(row=r, column=col_idx)
                if pd.notna(cell.value):
                    cell.number_format = '0.00'

######### Main report #########
def generate_sov_report(fasta_dir, perl_script_path, output_excel, Q=3):
    # Discover pairs
    pairs = find_fasta_pairs(fasta_dir)
    if not pairs:
        print("No valid pairs found — aborting.")
        return pd.DataFrame()

    if Q == 3:
        metric_cols = Q3_COLUMN_ORDER
        class_order = Q3_CLASS_ORDER
    else:
        metric_cols = Q8_COLUMN_ORDER
        class_order = Q8_CLASS_ORDER

    # Run every pair in parallel
    print(f"Spawning {len(pairs)} parallel thread(s)...\n")
    raw = {}
    with ThreadPoolExecutor(max_workers=len(pairs)) as executor:
        futures = {
            executor.submit(run_one_pair, base, p['pred'], p['true'], perl_script_path): base
            for base, p in pairs.items()
        }
        for future in as_completed(futures):
            model_name, results = future.result()
            raw[model_name] = results

    # Build one row per model (file name), fixed column order 
    rows = []
    for model_name in sorted(raw.keys()):
        results = raw[model_name]
        if results is None:
            print(f"  {model_name}: no result")
            continue
        rows.append({
            'Model': model_name,
            **{col: results.get(col) for col in metric_cols}
        })

    df = pd.DataFrame(rows)
    if df.empty:
        print("No data collected — check FASTA paths.")
        return pd.DataFrame()

    df = df[['Model'] + metric_cols]

    # Multiply to %
    df[metric_cols] = df[metric_cols].apply(
        lambda col: pd.to_numeric(col, errors='coerce') * 100
    )

    model_order  = list(df['Model'])
    model_colors = {
        m: MODEL_COLOR_PALETTE[i % len(MODEL_COLOR_PALETTE)]
        for i, m in enumerate(model_order)
    }

    # Write Excel
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='SOV_Results', index=False)
        ws   = writer.sheets['SOV_Results']
        cols = list(df.columns)

        _style_header_row(ws, cols, class_order)

        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
            row_fill = PatternFill("solid", fgColor=model_colors.get(row_data['Model'], "FFFFFF"))
            for col_idx in range(1, len(cols) + 1):
                cell           = ws.cell(row=row_idx, column=col_idx)
                cell.fill      = row_fill
                cell.font      = DARK_FONT
                cell.alignment = CENTER
                cell.border    = thin_border

        _apply_number_format(ws, cols, metric_cols, len(df))

        # Gold highlight — best per metric across all rows
        if len(df) >= 2:
            for col_name in metric_cols:
                col_idx = cols.index(col_name) + 1
                values  = {
                    r: ws.cell(row=r, column=col_idx).value
                    for r in range(2, len(df) + 2)
                    if ws.cell(row=r, column=col_idx).value is not None
                }
                if values:
                    best_row  = max(values, key=values.get)
                    best_cell = ws.cell(row=best_row, column=col_idx)
                    best_cell.fill = BEST_FILL
                    best_cell.font = Font(bold=True, color="1A1A2E")

        ws.freeze_panes = 'B2'
        _auto_col_widths(ws, cols, len(df))

    print(f"\n{'='*60}")
    print(f"Excel saved: {output_excel}")
    print(f"{'='*60}\n")
    print(df.to_string(index=False))

    return df

######### Usage #########
if __name__ == "__main__":

    perl_script_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        'SOV_refine', 'SOV_refine.pl'
    )
    fasta_dir = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        'SOV_refine', 'predictions'
    )

    output_excel = 'results.xlsx'

    df = generate_sov_report(
        fasta_dir        = fasta_dir,
        perl_script_path = perl_script_path,
        output_excel     = output_excel,
        Q                = 3,
    )
